import customtkinter
from customtkinter import CTkTextbox
import tkinter.messagebox as messagebox
import selenium
import requests
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook 
import os
import pandas as pd

customtkinter.set_appearance_mode("system")
customtkinter.set_default_color_theme("dark-blue")

root = customtkinter.CTk()
root.title("Movie Finder")
root.geometry("500x600")

title = year = runtime = genre = plot = first_rating = ""

def on_click():
    global title, year, runtime, genre, plot, first_rating
    movieTitle = entryBox.get()
    ##### using OMDb API
    url = f"http://www.omdbapi.com/?t={movieTitle}&apikey=7f56c7e2"
    response = requests.get(url)
    data = response.json()

    text_widget.configure(state="normal")
    title = data.get('Title', 'No info')
    year = data.get('Year', 'No info')
    runtime = data.get('Runtime', 'No info')
    genre = data.get('Genre', 'No info')
    plot = data.get('Plot', 'No info')
    ratings = data.get('Ratings', [{'Value': 'No info'}])
    first_rating = ratings[0]['Value'] if ratings else 'No info'

    text_widget.delete('1.0', 'end')
    text_widget.insert('end', f"Title: {title}\n\nYear: {year}\n\nRuntime: {runtime}\n\nGenre: {genre}\n\nPlot: {plot}\n\nRatings: {first_rating}")
    text_widget.configure(state="disabled")
    
    if title=='No info':
        messagebox.showerror("Error", "This movie doesn't exist")
        TrailerButton.pack_forget()
        ExcelWriteButton.pack_forget()
    else: 
        TrailerButton.pack(pady=12, padx=10)
        ExcelWriteButton.pack(pady=12, padx=10)

    root.focus()

def on_enter(event):
    on_click()

def open_trailer():
    global title, year
    service = Service()
    option = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=service, options=option)

    url = "https://www.imdb.com/"
    driver.get(url)

    time.sleep(1)

    find=driver.find_element(By.CLASS_NAME, "icb-btn")
    find.click()

    time.sleep(1)

    find=driver.find_element(By.ID, "suggestion-search")
    find.send_keys(title, " ", year)
    time.sleep(2)
    find=driver.find_element(By.CLASS_NAME, "ipc-icon.ipc-icon--play-circle-outline-inline.ipc-lockup-overlay__icon")
    find.click()

    input()

def save_to_excel():
    global title, year, runtime, genre, plot, first_rating
    #print(title, year, runtime, genre, first_rating)
    filepath = "movies.xlsx"
    fullpath = os.path.join(os.path.dirname(__file__), filepath)

    if os.path.isfile(fullpath):
        wb = load_workbook(fullpath)
    else:
        wb = Workbook()
        wb.save(fullpath)
        wb = load_workbook(fullpath)

    ws = wb.active
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 35
    ws['A1'] = "Title"
    ws['B1'] = "Runtime"
    ws['C1'] = "Rating"
    ws['D1'] = "Genre"
    for row in range(2, ws.max_row + 2):  # Добавляем +2, чтобы проверить строку после последней строки с записью
        cell_value = ws['A' + str(row)].value
        if cell_value is None or str(cell_value).strip() == "":
            print(title)
            ws['A' + str(row)] = title
            ws['B' + str(row)] = runtime
            ws['C' + str(row)] = first_rating
            ws['D' + str(row)] = genre
            break

    wb.save("movies.xlsx")
    wb.close()



frame = customtkinter.CTkFrame(master=root)
frame.pack(pady=0, padx=0, fill="both", expand=True)

label = customtkinter.CTkLabel(master=frame, text="Finding information about movie or series", font=("TkTextFont", 20))
label.pack(pady=12, padx=10)
label = customtkinter.CTkLabel(master=frame, text="You should write full name of movie in english", font=("TkTextFont", 14))
label.pack(pady=12, padx=10)

entryBox = customtkinter.CTkEntry(master=frame, placeholder_text="Enter movie or series name", font=("TkTextFont", 16), width=208, height=20)
entryBox.bind('<Return>', on_enter)
entryBox.pack(pady=12, padx=20)

button = customtkinter.CTkButton(master=frame, text="Find", command=on_click)
button.pack(pady=12, padx=10)

text_widget = CTkTextbox(master=frame, width=500, height=250)
text_widget.configure(state="disabled")
text_widget.pack(pady=12, padx=10)

TrailerButton = customtkinter.CTkButton(master=frame, text="Open trailer", command=open_trailer)
TrailerButton.pack_forget()
ExcelWriteButton = customtkinter.CTkButton(master=frame, text="Save to Excel document", command=save_to_excel)
ExcelWriteButton.pack_forget()

### TO_DO: write info about movie in excel document to watch it later

root.mainloop()
